import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from models import OMRSheet


def auto_fit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                value = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(value))
            except:
                pass

        ws.column_dimensions[col_letter].width = max(max_len + 2, 12)


def build_excel(sheet_ids=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "OMR Results"

    query = OMRSheet.query

    if sheet_ids:
        query = query.filter(OMRSheet.id.in_(sheet_ids))

    sheets = query.order_by(OMRSheet.percentage.desc()).all()

    if not sheets:
        return None

    first_sheet_answers = sheets[0].answers or {}
    total_questions = len(first_sheet_answers)

    headers = [
        "Sl.No",
        "File Name",
        "Name",
        "Roll Number",
        "Class",
        "Section",
        "Stream",
        "Set",
        "Subject Code",
        "Admission No"
    ]
    
    sl_no = 1

    for i in range(1, total_questions + 1):
        headers.append(f"Q{str(i).zfill(3)}")

    headers += ["Correct", "Wrong", "Percentage"]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    for sheet in sheets:
        row = [
            sl_no,
            sheet.result_file,
            sheet.name,
            sheet.roll_number,
            sheet.class_name,
            sheet.section,
            sheet.stream,
            sheet.set_number,
            sheet.subject_code,
            sheet.admission_number
        ]

        answers_json = sheet.answers or {}

        for i in range(1, total_questions + 1):
            q_no = f"Q{str(i).zfill(3)}"

            ans_data = answers_json.get(q_no)

            if ans_data:
                row.append(ans_data.get("selected"))
            else:
                row.append("")

        row += [
            sheet.correct_answers,
            sheet.wrong_answers,
            round(sheet.percentage, 2)
        ]

        ws.append(row)
        sl_no+=1

    auto_fit_columns(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"omr_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return output, filename